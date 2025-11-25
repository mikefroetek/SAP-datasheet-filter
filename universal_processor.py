"""
Universal Excel Level Processor
Handles unlimited levels (1 to 1000+) with dynamic hierarchical structure.
Follows BOM Template Aptiv pattern for any depth of hierarchy.
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os
import re
from datetime import datetime

def process_aptive_to_bom_universal(aptive_file_path, bom_template_path, output_path=None):
    """
    Process Aptive Excel file with unlimited levels following hierarchical BOM pattern.
    
    Dynamic Pattern:
    - Level N: Gets Material row (Component empty) 
    - Level N+1: Appears as Component under Level N, then gets own Material row
    - Level N+2: Appears as Component under Level N+1, and so on...
    - Can handle levels 1 to 1000+
    """
    
    print("🔄 Starting Universal Excel processing (unlimited levels)...")
    print("📋 Following BOM Template Aptiv hierarchical pattern")
    
    # Generate output path if not provided
    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_path = os.path.dirname(bom_template_path)
        output_path = os.path.join(base_path, f"BOM_Universal_{timestamp}.xlsx")
    
    try:
        # Step 1: Read Aptive file
        print(f"📖 Reading Aptive file: {os.path.basename(aptive_file_path)}")
        
        if aptive_file_path.endswith('.xlsb'):
            df = pd.read_excel(aptive_file_path, engine='pyxlsb')
        else:
            df = pd.read_excel(aptive_file_path)
        
        print(f"✓ Read {len(df)} rows from Aptive file")
        
        # Analyze level distribution
        levels_found = set()
        for i in range(1, len(df)):
            if not pd.isna(df.iloc[i, 0]):
                levels_found.add(int(df.iloc[i, 0]))
        
        max_level = max(levels_found) if levels_found else 1
        print(f"📊 Levels detected: {sorted(levels_found)} (Max: {max_level})")
        
        # Step 2: Load BOM template and preserve rows 1-8
        print("📋 Loading BOM template (preserving rows 1-8)...")
        workbook = load_workbook(bom_template_path)
        worksheet = workbook.active
        
        current_row = 9  # START FROM ROW 9 - DO NOT MODIFY ROWS 1-8
        
        print("🔄 Processing unlimited-level hierarchical structure...")
        
        # Step 3: Process data with recursive hierarchy handling
        processed_groups = process_hierarchy_recursive(df, worksheet, current_row, max_level)
        
        # Save the output file
        workbook.save(output_path)
        
        print(f"✅ Processing completed successfully!")
        print(f"📊 Processed {processed_groups} hierarchy groups")
        print(f"📁 Output saved to: {output_path}")
        print(f"🔒 Template rows 1-8 preserved, data starts from row 9")
        print(f"🚀 Handled levels 1 to {max_level}")
        
        return output_path
        
    except Exception as e:
        print(f"❌ Error during processing: {str(e)}")
        raise

def process_hierarchy_recursive(df, worksheet, start_row, max_level):
    """
    Recursively process hierarchical data with unlimited levels
    """
    current_row = start_row
    processed_groups = 0
    
    i = 1  # Start from row 1 (skip header row 0)
    while i < len(df):
        row = df.iloc[i]
        
        # Skip empty rows
        if pd.isna(row.iloc[0]):
            i += 1
            continue
        
        level = int(row.iloc[0])
        
        if level == 1:
            # Found Level 1 - start new hierarchy group
            processed_groups += 1
            art_nr1 = str(int(row.iloc[1])) if not pd.isna(row.iloc[1]) else ""
            
            print(f"  📦 Group {processed_groups}: Level 1 - {art_nr1}")
            
            # Level 1: Create Material row (Component empty)
            worksheet.cell(row=current_row, column=5, value=art_nr1)
            worksheet.cell(row=current_row, column=14, value="")
            worksheet.cell(row=current_row, column=15, value="")
            worksheet.cell(row=current_row, column=16, value="")
            current_row += 1
            
            # Process all child levels recursively
            current_row, next_index = process_child_levels(
                df, worksheet, i, current_row, level, art_nr1, max_level
            )
            
            i = next_index
        else:
            i += 1  # Skip non-Level 1 items (processed as components)
    
    return processed_groups

def process_child_levels(df, worksheet, parent_index, current_row, parent_level, parent_art_nr, max_level):
    """
    Recursively process child levels under a parent
    """
    child_level = parent_level + 1
    child_items = []
    
    # Find all direct children of current parent
    i = parent_index + 1
    while i < len(df) and not pd.isna(df.iloc[i, 0]):
        row_level = int(df.iloc[i, 0])
        
        if row_level == child_level:
            child_items.append(i)
        elif row_level <= parent_level:
            break  # Found item at same or higher level, stop looking for children
        
        i += 1
    
    # Process each child as component under parent
    for child_index in child_items:
        child_row = df.iloc[child_index]
        child_art_nr = str(int(child_row.iloc[1])) if not pd.isna(child_row.iloc[1]) else ""
        child_unit = str(child_row.iloc[3]).upper() if not pd.isna(child_row.iloc[3]) else ""
        child_quantity = child_row.iloc[4] if not pd.isna(child_row.iloc[4]) else ""
        
        indent = "  " * child_level
        print(f"{indent}🔗 Level {child_level} under {parent_art_nr}: {child_art_nr}")
        
        # Child as Component under Parent Material
        worksheet.cell(row=current_row, column=5, value=parent_art_nr)      # Parent Material (E)
        worksheet.cell(row=current_row, column=14, value=child_art_nr)      # Child Component (N)
        worksheet.cell(row=current_row, column=15, value=child_quantity)    # Quantity (O)
        worksheet.cell(row=current_row, column=16, value=child_unit)        # Unit (P)
        current_row += 1
    
    # Now each child becomes a Material with its own children (if any)
    for child_index in child_items:
        child_row = df.iloc[child_index]
        child_art_nr = str(int(child_row.iloc[1])) if not pd.isna(child_row.iloc[1]) else ""
        
        # Check if this child has any children of its own
        has_grandchildren = False
        j = child_index + 1
        while j < len(df) and not pd.isna(df.iloc[j, 0]):
            if int(df.iloc[j, 0]) == child_level + 1:
                has_grandchildren = True
                break
            elif int(df.iloc[j, 0]) <= child_level:
                break
            j += 1
        
        if has_grandchildren:
            # Child gets its own Material row (for grandchildren)
            worksheet.cell(row=current_row, column=5, value=child_art_nr)   # Child as Material (E)
            worksheet.cell(row=current_row, column=14, value="")            # Component empty (N)
            worksheet.cell(row=current_row, column=15, value="")            # Quantity empty (O)
            worksheet.cell(row=current_row, column=16, value="")            # Unit empty (P)
            current_row += 1
            
            # Recursively process grandchildren
            current_row, _ = process_child_levels(
                df, worksheet, child_index, current_row, child_level, child_art_nr, max_level
            )
    
    # Return next unprocessed index
    next_index = parent_index + 1
    while next_index < len(df) and not pd.isna(df.iloc[next_index, 0]):
        if int(df.iloc[next_index, 0]) <= parent_level:
            break
        next_index += 1
    
    return current_row, next_index

def main():
    """Main function"""
    
    # File paths
    aptive_file = r"d:\SAP\Aptive_BOOM_v01.xlsb"
    bom_template = r"d:\SAP\BOM_Template.xlsx"
    
    # Check if files exist
    if not os.path.exists(aptive_file):
        print(f"❌ Aptive file not found: {aptive_file}")
        return
        
    if not os.path.exists(bom_template):
        print(f"❌ BOM Template file not found: {bom_template}")
        return
    
    # Process the files
    try:
        output_file = process_aptive_to_bom_universal(aptive_file, bom_template)
        print(f"\n🎉 Success! Universal processor handles unlimited levels!")
        print(f"📁 Check your output file: {os.path.basename(output_file)}")
        
    except Exception as e:
        print(f"\n💥 Failed to process files: {str(e)}")

if __name__ == "__main__":
    main()