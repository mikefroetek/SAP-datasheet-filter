"""
Sequential Hierarchical Excel Level Processor
Handles sequential parent-child relationships where each level relates to the previous one.
When a level repeats, it starts a new branch.

Example:
Szint: 1, 2, 3, 4, 3, 4, 4
- 1 → 2 → 3 → 4
- When 3 repeats, it starts new branch: 3 → 4, 4
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os
import re
from datetime import datetime

def process_aptive_sequential(aptive_file_path, bom_template_path, output_path=None):
    """
    Process Aptive Excel file with sequential hierarchical logic.
    """
    
    print("🔄 Starting Sequential Hierarchical Excel processing...")
    print("📋 Each level relates to the immediately previous level")
    
    # Generate output path if not provided
    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_path = os.path.dirname(bom_template_path)
        output_path = os.path.join(base_path, f"BOM_Sequential_{timestamp}.xlsx")
    
    try:
        # Step 1: Read Aptive file
        print(f"📖 Reading Aptive file: {os.path.basename(aptive_file_path)}")
        
        if aptive_file_path.endswith('.xlsb'):
            df = pd.read_excel(aptive_file_path, engine='pyxlsb')
        else:
            df = pd.read_excel(aptive_file_path)
        
        print(f"✓ Read {len(df)} rows from Aptive file")
        
        # Analyze level distribution
        levels_found = set()
        for i in range(1, len(df)):
            if not pd.isna(df.iloc[i, 0]):
                levels_found.add(int(df.iloc[i, 0]))
        
        max_level = max(levels_found) if levels_found else 1
        print(f"📊 Levels detected: {sorted(levels_found)} (Max: {max_level})")
        
        # Step 2: Load BOM template and preserve rows 1-8
        print("📋 Loading BOM template (preserving rows 1-8)...")
        workbook = load_workbook(bom_template_path)
        worksheet = workbook.active
        
        current_row = 9  # START FROM ROW 9 - DO NOT MODIFY ROWS 1-8
        
        print("🔄 Processing sequential hierarchical structure...")
        
        # Step 3: Process data with sequential logic
        current_row = process_sequential_hierarchy(df, worksheet, current_row)
        
        # Save the output file
        workbook.save(output_path)
        
        print(f"✅ Processing completed successfully!")
        print(f"📁 Output saved to: {output_path}")
        print(f"🔒 Template rows 1-8 preserved, data starts from row 9")
        print(f"🔗 Sequential parent-child relationships processed")
        
        return output_path
        
    except Exception as e:
        print(f"❌ Error during processing: {str(e)}")
        raise

def process_sequential_hierarchy(df, worksheet, start_row):
    """
    Process data with specific ordering: all same-level components under parent first,
    then their children in sequence
    """
    current_row = start_row
    
    print("🔍 Processing sequential relationships with ordered level handling...")
    
    # First pass: build hierarchy structure
    hierarchy = build_hierarchy_structure(df)
    
    # Second pass: write to BOM with proper ordering
    current_row = write_hierarchy_ordered(hierarchy, worksheet, current_row)
    
    return current_row

def build_hierarchy_structure(df):
    """
    Build hierarchy structure from the dataframe
    """
    items = []
    
    for i in range(1, len(df)):  # Skip header row
        row = df.iloc[i]
        
        # Skip empty rows
        if pd.isna(row.iloc[0]):
            continue
            
        level = int(row.iloc[0])
        art_nr = str(int(row.iloc[1])) if not pd.isna(row.iloc[1]) else ""
        unit = str(row.iloc[3]).upper() if not pd.isna(row.iloc[3]) else ""
        quantity = row.iloc[4] if not pd.isna(row.iloc[4]) else ""
        
        items.append({
            'level': level,
            'art_nr': art_nr,
            'unit': unit,
            'quantity': quantity,
            'children': []
        })
    
    # Build parent-child relationships
    for i, item in enumerate(items):
        current_level = item['level']
        
        # Find children (items at current_level + 1 that come after this item)
        j = i + 1
        while j < len(items):
            if items[j]['level'] == current_level + 1:
                item['children'].append(items[j])
            elif items[j]['level'] <= current_level:
                break  # Found item at same or higher level, stop looking
            j += 1
    
    # Return only Level 1 items (roots)
    return [item for item in items if item['level'] == 1]

def write_hierarchy_ordered(hierarchy, worksheet, start_row):
    """
    Write hierarchy with proper ordering: all components of same level first,
    then their children. Also fills column L (Item number) with sequential numbering.
    """
    current_row = start_row
    
    for level1_item in hierarchy:
        print(f"  📦 Processing Level 1: {level1_item['art_nr']}")
        
        # Write Level 1 as Material
        worksheet.cell(row=current_row, column=5, value=level1_item['art_nr'])
        worksheet.cell(row=current_row, column=12, value="")  # Column L - empty for Material
        worksheet.cell(row=current_row, column=14, value="")
        worksheet.cell(row=current_row, column=15, value="")
        worksheet.cell(row=current_row, column=16, value="")
        current_row += 1
        
        # Write all Level 2 components under Level 1
        level2_items = level1_item['children']
        item_number = 10  # Start with 0010
        for level2_item in level2_items:
            print(f"    🔗 Level 1 → Level 2: {level1_item['art_nr']} → {level2_item['art_nr']}")
            worksheet.cell(row=current_row, column=5, value=level1_item['art_nr'])
            worksheet.cell(row=current_row, column=12, value=f"{item_number:04d}")  # Column L - sequential numbering
            worksheet.cell(row=current_row, column=14, value=level2_item['art_nr'])
            worksheet.cell(row=current_row, column=15, value=level2_item['quantity'])
            worksheet.cell(row=current_row, column=16, value=level2_item['unit'])
            current_row += 1
            item_number += 10  # Increment by 10 (0010, 0020, 0030, etc.)
        
        # Now process each Level 2 and its children
        for level2_item in level2_items:
            if level2_item['children']:  # Has Level 3 children
                print(f"    📦 Level 2 as Material: {level2_item['art_nr']}")
                
                # Write Level 2 as Material
                worksheet.cell(row=current_row, column=5, value=level2_item['art_nr'])
                worksheet.cell(row=current_row, column=12, value="")  # Column L - empty for Material
                worksheet.cell(row=current_row, column=14, value="")
                worksheet.cell(row=current_row, column=15, value="")
                worksheet.cell(row=current_row, column=16, value="")
                current_row += 1
                
                # Write all Level 3 components under Level 2
                level3_items = level2_item['children']
                item_number = 10  # Reset numbering for each new Material
                for level3_item in level3_items:
                    print(f"      🔗 Level 2 → Level 3: {level2_item['art_nr']} → {level3_item['art_nr']}")
                    worksheet.cell(row=current_row, column=5, value=level2_item['art_nr'])
                    worksheet.cell(row=current_row, column=12, value=f"{item_number:04d}")  # Column L - sequential numbering
                    worksheet.cell(row=current_row, column=14, value=level3_item['art_nr'])
                    worksheet.cell(row=current_row, column=15, value=level3_item['quantity'])
                    worksheet.cell(row=current_row, column=16, value=level3_item['unit'])
                    current_row += 1
                    item_number += 10  # Increment by 10
                
                # Now process each Level 3 and its children
                for level3_item in level3_items:
                    if level3_item['children']:  # Has Level 4 children
                        print(f"      � Level 3 as Material: {level3_item['art_nr']}")
                        
                        # Write Level 3 as Material
                        worksheet.cell(row=current_row, column=5, value=level3_item['art_nr'])
                        worksheet.cell(row=current_row, column=12, value="")  # Column L - empty for Material
                        worksheet.cell(row=current_row, column=14, value="")
                        worksheet.cell(row=current_row, column=15, value="")
                        worksheet.cell(row=current_row, column=16, value="")
                        current_row += 1
                        
                        # Write all Level 4 components under Level 3
                        level4_items = level3_item['children']
                        item_number = 10  # Reset numbering for each new Material
                        for level4_item in level4_items:
                            print(f"        � Level 3 → Level 4: {level3_item['art_nr']} → {level4_item['art_nr']}")
                            worksheet.cell(row=current_row, column=5, value=level3_item['art_nr'])
                            worksheet.cell(row=current_row, column=12, value=f"{item_number:04d}")  # Column L - sequential numbering
                            worksheet.cell(row=current_row, column=14, value=level4_item['art_nr'])
                            worksheet.cell(row=current_row, column=15, value=level4_item['quantity'])
                            worksheet.cell(row=current_row, column=16, value=level4_item['unit'])
                            current_row += 1
                            item_number += 10  # Increment by 10
                            
                            # Continue pattern for deeper levels if needed
                            current_row = write_deeper_levels(level4_item, worksheet, current_row, 4)
    
    return current_row

def write_deeper_levels(parent_item, worksheet, current_row, parent_level):
    """
    Recursively write deeper levels (5, 6, 7, ... up to 1000)
    """
    if not parent_item['children']:
        return current_row
    
    # Write parent as Material if it has children
    worksheet.cell(row=current_row, column=5, value=parent_item['art_nr'])
    worksheet.cell(row=current_row, column=12, value="")  # Column L - empty for Material
    worksheet.cell(row=current_row, column=14, value="")
    worksheet.cell(row=current_row, column=15, value="")
    worksheet.cell(row=current_row, column=16, value="")
    current_row += 1
    
    # Write all children as components
    item_number = 10  # Reset numbering for each new Material
    for child_item in parent_item['children']:
        indent = "  " * (parent_level - 2)
        print(f"{indent}🔗 Level {parent_level} → Level {parent_level + 1}: {parent_item['art_nr']} → {child_item['art_nr']}")
        worksheet.cell(row=current_row, column=5, value=parent_item['art_nr'])
        worksheet.cell(row=current_row, column=12, value=f"{item_number:04d}")  # Column L - sequential numbering
        worksheet.cell(row=current_row, column=14, value=child_item['art_nr'])
        worksheet.cell(row=current_row, column=15, value=child_item['quantity'])
        worksheet.cell(row=current_row, column=16, value=child_item['unit'])
        current_row += 1
        item_number += 10  # Increment by 10
    
    # Process each child's children
    for child_item in parent_item['children']:
        current_row = write_deeper_levels(child_item, worksheet, current_row, parent_level + 1)
    
    return current_row

def test_sequential_logic():
    """
    Test the sequential logic with the example provided
    """
    print("🧪 Testing sequential logic with ordered example:")
    print("Levels: 1, 2, 3, 4, 3, 4, 4")
    print("Expected ORDERED relationships:")
    print("  1 → 2")
    print("  2 → 3 (first)")
    print("  2 → 3 (second) [both 3s go under same 2 FIRST]")
    print("  3 (first) → 4")
    print("  3 (second) → 4, 4 [both 4s go under second 3]")
    print()
    print("Expected BOM output order:")
    print("  1. Material: A001 (Level 1)")
    print("  2. A001 → A002 (Level 2)")
    print("  3. Material: A002 (Level 2)")
    print("  4. A002 → A003 (Level 3 first)")
    print("  5. A002 → A005 (Level 3 second)")
    print("  6. Material: A003 (Level 3 first)")
    print("  7. A003 → A004 (Level 4)")
    print("  8. Material: A005 (Level 3 second)")
    print("  9. A005 → A006 (Level 4 first)")
    print(" 10. A005 → A007 (Level 4 second)")
    print()

def main():
    """Main function"""
    
    # Test the logic first
    test_sequential_logic()
    
    # File paths
    aptive_file = r"d:\SAP\Aptive_BOOM_v01.xlsb"
    bom_template = r"d:\SAP\BOM_Template.xlsx"
    
    # Check if files exist
    if not os.path.exists(aptive_file):
        print(f"❌ Aptive file not found: {aptive_file}")
        return
        
    if not os.path.exists(bom_template):
        print(f"❌ BOM Template file not found: {bom_template}")
        return
    
    # Process the files
    try:
        output_file = process_aptive_sequential(aptive_file, bom_template)
        print(f"\n🎉 Success! Sequential hierarchical processor completed!")
        print(f"📁 Check your output file: {os.path.basename(output_file)}")
        
    except Exception as e:
        print(f"\n💥 Failed to process files: {str(e)}")

if __name__ == "__main__":
    main()