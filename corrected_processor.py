"""
Corrected Excel Level Processor
Follows the exact BOM Template Aptiv pattern with hierarchical structure.
Data insertion starts from row 9, preserving template rows 1-8.
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os
import re
from datetime import datetime

def process_aptive_to_bom_correct(aptive_file_path, bom_template_path, output_path=None):
    """
    Process Aptive Excel file following the exact BOM Template Aptiv hierarchical pattern.
    
    Pattern from BOM Template Aptiv:
    - Level 1: Gets Material row (Component empty) 
    - Level 2: Appears as Component under Level 1, then gets own Material row
    - Level 3: Appears as Component under Level 2, and so on
    - Data starts from row 9, preserving template structure rows 1-8
    """
    
    print("🔄 Starting corrected Excel processing...")
    print("📋 Following BOM Template Aptiv hierarchical pattern")
    
    # Generate output path if not provided
    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_path = os.path.dirname(bom_template_path)
        output_path = os.path.join(base_path, f"BOM_Corrected_{timestamp}.xlsx")
    
    try:
        # Step 1: Read Aptive file
        print(f"📖 Reading Aptive file: {os.path.basename(aptive_file_path)}")
        
        if aptive_file_path.endswith('.xlsb'):
            df = pd.read_excel(aptive_file_path, engine='pyxlsb')
        else:
            df = pd.read_excel(aptive_file_path)
        
        print(f"✓ Read {len(df)} rows from Aptive file")
        
        # Step 2: Load BOM template and preserve rows 1-8
        print("📋 Loading BOM template (preserving rows 1-8)...")
        workbook = load_workbook(bom_template_path)
        worksheet = workbook.active
        
        current_row = 9  # START FROM ROW 9 - DO NOT MODIFY ROWS 1-8
        
        print("🔄 Processing hierarchical structure...")
        
        # Step 3: Process data maintaining parent-child relationships
        i = 1  # Start from row 1 (skip header row 0)
        processed_groups = 0
        
        while i < len(df):
            row = df.iloc[i]
            
            # Skip empty rows
            if pd.isna(row.iloc[0]):
                i += 1
                continue
            
            level = int(row.iloc[0])
            
            if level == 1:
                # Found a Level 1 item - start a new hierarchy group
                processed_groups += 1
                art_nr1 = str(int(row.iloc[1])) if not pd.isna(row.iloc[1]) else ""
                
                print(f"  📦 Group {processed_groups}: Level 1 - {art_nr1}")
                
                # Level 1: Create Material row (Component empty)
                worksheet.cell(row=current_row, column=5, value=art_nr1)    # Material (E)
                worksheet.cell(row=current_row, column=14, value="")        # Component empty (N)
                worksheet.cell(row=current_row, column=15, value="")        # Quantity empty (O)
                worksheet.cell(row=current_row, column=16, value="")        # Unit empty (P)
                current_row += 1
                
                # Look for Level 2 items that belong to this Level 1
                j = i + 1
                while j < len(df) and not pd.isna(df.iloc[j, 0]) and int(df.iloc[j, 0]) >= 2:
                    row2 = df.iloc[j]
                    level2 = int(row2.iloc[0])
                    
                    if level2 == 2:
                        art_nr2 = str(int(row2.iloc[1])) if not pd.isna(row2.iloc[1]) else ""
                        unit2 = str(row2.iloc[3]).upper() if not pd.isna(row2.iloc[3]) else ""
                        quantity2 = row2.iloc[4] if not pd.isna(row2.iloc[4]) else ""
                        
                        print(f"    🔗 Level 2 under {art_nr1}: {art_nr2}")
                        
                        # Level 2 as Component under Level 1 Material
                        worksheet.cell(row=current_row, column=5, value=art_nr1)       # Level 1 Material (E)
                        worksheet.cell(row=current_row, column=14, value=art_nr2)      # Level 2 Component (N)
                        worksheet.cell(row=current_row, column=15, value=quantity2)    # Quantity (O)
                        worksheet.cell(row=current_row, column=16, value=unit2)        # Unit (P)
                        current_row += 1
                        
                        # Level 2 also gets its own Material row (for potential Level 3)
                        worksheet.cell(row=current_row, column=5, value=art_nr2)       # Level 2 as Material (E)
                        worksheet.cell(row=current_row, column=14, value="")           # Component empty (N)
                        worksheet.cell(row=current_row, column=15, value="")           # Quantity empty (O)
                        worksheet.cell(row=current_row, column=16, value="")           # Unit empty (P)
                        current_row += 1
                        
                        # Collect all Level 3 items under this Level 2
                        level3_items = []
                        k = j + 1
                        while k < len(df) and not pd.isna(df.iloc[k, 0]) and int(df.iloc[k, 0]) >= 3:
                            if int(df.iloc[k, 0]) == 3:
                                level3_items.append(k)
                            k += 1
                        
                        # Process each Level 3 item as component under Level 2
                        for k in level3_items:
                            row3 = df.iloc[k]
                            art_nr3 = str(int(row3.iloc[1])) if not pd.isna(row3.iloc[1]) else ""
                            unit3 = str(row3.iloc[3]).upper() if not pd.isna(row3.iloc[3]) else ""
                            quantity3 = row3.iloc[4] if not pd.isna(row3.iloc[4]) else ""
                            
                            print(f"      🔗 Level 3 under {art_nr2}: {art_nr3}")
                            
                            # Level 3 as Component under Level 2 Material
                            worksheet.cell(row=current_row, column=5, value=art_nr2)       # Level 2 Material (E)
                            worksheet.cell(row=current_row, column=14, value=art_nr3)      # Level 3 Component (N)
                            worksheet.cell(row=current_row, column=15, value=quantity3)    # Quantity (O)
                            worksheet.cell(row=current_row, column=16, value=unit3)        # Unit (P)
                            current_row += 1
                        
                        # Now each Level 3 gets its own Material section with Level 4 components
                        for k in level3_items:
                            row3 = df.iloc[k]
                            art_nr3 = str(int(row3.iloc[1])) if not pd.isna(row3.iloc[1]) else ""
                            
                            # Level 3 gets its own Material row
                            worksheet.cell(row=current_row, column=5, value=art_nr3)       # Level 3 as Material (E)
                            worksheet.cell(row=current_row, column=14, value="")           # Component empty (N)
                            worksheet.cell(row=current_row, column=15, value="")           # Quantity empty (O)
                            worksheet.cell(row=current_row, column=16, value="")           # Unit empty (P)
                            current_row += 1
                            
                            # Find all Level 4 items under this Level 3
                            l = k + 1
                            while l < len(df) and not pd.isna(df.iloc[l, 0]) and int(df.iloc[l, 0]) == 4:
                                row4 = df.iloc[l]
                                art_nr4 = str(int(row4.iloc[1])) if not pd.isna(row4.iloc[1]) else ""
                                unit4 = str(row4.iloc[3]).upper() if not pd.isna(row4.iloc[3]) else ""
                                quantity4 = row4.iloc[4] if not pd.isna(row4.iloc[4]) else ""
                                
                                print(f"        🔗 Level 4 under {art_nr3}: {art_nr4}")
                                
                                # Level 4 as Component under Level 3 Material
                                worksheet.cell(row=current_row, column=5, value=art_nr3)       # Level 3 Material (E)
                                worksheet.cell(row=current_row, column=14, value=art_nr4)      # Level 4 Component (N)
                                worksheet.cell(row=current_row, column=15, value=quantity4)    # Quantity (O)
                                worksheet.cell(row=current_row, column=16, value=unit4)        # Unit (P)
                                current_row += 1
                                
                                l += 1
                        
                        j = k  # Skip processed Level 3 items
                    else:
                        break  # Not Level 2, exit loop
                    
                    j += 1
                
                i = j  # Skip processed Level 2+ items
            else:
                i += 1  # Skip non-Level 1 items (they're processed as components)
        
        # Save the output file
        workbook.save(output_path)
        
        print(f"✅ Processing completed successfully!")
        print(f"📊 Processed {processed_groups} hierarchy groups")
        print(f"📁 Output saved to: {output_path}")
        print(f"🔒 Template rows 1-8 preserved, data starts from row 9")
        
        return output_path
        
    except Exception as e:
        print(f"❌ Error during processing: {str(e)}")
        raise

def main():
    """Main function"""
    
    # File paths
    aptive_file = r"d:\SAP\Aptive_BOOM_v01.xlsb"
    bom_template = r"d:\SAP\BOM_Template.xlsx"
    
    # Check if files exist
    if not os.path.exists(aptive_file):
        print(f"❌ Aptive file not found: {aptive_file}")
        return
        
    if not os.path.exists(bom_template):
        print(f"❌ BOM Template file not found: {bom_template}")
        return
    
    # Process the files
    try:
        output_file = process_aptive_to_bom_correct(aptive_file, bom_template)
        print(f"\n🎉 Success! Check your output file: {os.path.basename(output_file)}")
        
    except Exception as e:
        print(f"\n💥 Failed to process files: {str(e)}")

if __name__ == "__main__":
    main()