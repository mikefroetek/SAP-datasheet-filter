"""
Simple Excel Level Processor
A simplified version for processing Aptive BOOM files into BOM Template format.
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os
import re
from datetime import datetime

def process_aptive_to_bom(aptive_file_path, bom_template_path, output_path=None):
    """
    Process Aptive Excel file and write to BOM Template based on levels.
    
    Args:
        aptive_file_path (str): Path to Aptive BOOM Excel file
        bom_template_path (str): Path to BOM Template Excel file  
        output_path (str): Optional output path (auto-generated if None)
    
    Returns:
        str: Path to the created output file
    """
    
    print("🔄 Starting Excel processing...")
    
    # Generate output path if not provided
    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_path = os.path.dirname(bom_template_path)
        output_path = os.path.join(base_path, f"BOM_Processed_{timestamp}.xlsx")
    
    try:
        # Step 1: Read Aptive file
        print(f"📖 Reading Aptive file: {os.path.basename(aptive_file_path)}")
        
        # Handle .xlsb format
        if aptive_file_path.endswith('.xlsb'):
            df = pd.read_excel(aptive_file_path, engine='pyxlsb')
        else:
            df = pd.read_excel(aptive_file_path)
        
        print(f"✓ Read {len(df)} rows from Aptive file")
        
        # Step 2: Extract and organize data by levels
        print("🔄 Processing level-wise data...")
        
        # Get columns A, B, D, E (indices 0, 1, 3, 4)
        level_data = {}
        
        for index, row in df.iterrows():
            # Skip header row and empty rows
            if index == 0 or pd.isna(row.iloc[0]):
                continue
                
            # Extract level from column A (Szint)
            level_value = str(row.iloc[0]).strip()
            
            # Extract numeric level
            numbers = re.findall(r'\d+', level_value)
            if not numbers:
                continue
                
            level = int(numbers[0])
            
            # Extract data from columns B, D, E
            col_b = str(row.iloc[1]) if not pd.isna(row.iloc[1]) else ""
            col_d = str(row.iloc[3]) if not pd.isna(row.iloc[3]) else ""
            col_e = str(row.iloc[4]) if not pd.isna(row.iloc[4]) else ""
            
            # Store data by level
            if level not in level_data:
                level_data[level] = []
                
            level_data[level].append({
                'level': level,             # Numeric level
                'level_text': level_value,  # Original level text
                'col_b': col_b,
                'col_d': col_d,
                'col_e': col_e
            })
        
        # Print level summary
        for level in sorted(level_data.keys()):
            print(f"  Level {level}: {len(level_data[level])} items")
        
        # Step 3: Write to BOM Template
        print(f"✏️ Writing to BOM Template...")
        
        # Load BOM template
        workbook = load_workbook(bom_template_path)
        worksheet = workbook.active
        
        # Start writing from row 2 (assuming row 1 has headers)
        current_row = 2
        
        # Write data level by level
        for level in sorted(level_data.keys()):
            print(f"  Writing Level {level} data...")
            
            for item in level_data[level]:
                # Based on BOM Template Aptiv example:
                # Level 1: Art. Nr (Aptive col B) goes to Material column (BOM E)
                # Level 2+: Art. Nr (Aptive col B) goes to Component column (BOM N)
                # Quantities and units from Aptive cols D, E go to BOM cols O, P
                
                if item['level'] == 1:
                    # Level 1: Put Art. Nr in Material column (E)
                    worksheet.cell(row=current_row, column=5, value=item['col_b'])    # Column E (Material)
                    worksheet.cell(row=current_row, column=14, value='')              # Column N (Component) - empty for level 1
                else:
                    # Level 2+: Put Art. Nr in Component column (N)
                    worksheet.cell(row=current_row, column=5, value='')               # Column E (Material) - empty for level 2+
                    worksheet.cell(row=current_row, column=14, value=item['col_b'])   # Column N (Component)
                
                # Always put quantity and unit data
                worksheet.cell(row=current_row, column=15, value=item['col_d'])       # Column O (Quantity)
                worksheet.cell(row=current_row, column=16, value=item['col_e'])       # Column P (Unit)
                
                current_row += 1
            
            # Add blank row between levels
            current_row += 1        # Save the output file
        workbook.save(output_path)
        
        print(f"✅ Processing completed successfully!")
        print(f"📁 Output saved to: {output_path}")
        
        return output_path
        
    except Exception as e:
        print(f"❌ Error during processing: {str(e)}")
        raise

def main():
    """Main function - modify file paths as needed"""
    
    # File paths - UPDATE THESE AS NEEDED
    aptive_file = r"d:\SAP\Aptive_BOOM_v01.xlsb"
    bom_template = r"d:\SAP\BOM_Template.xlsx"
    
    # Check if files exist
    if not os.path.exists(aptive_file):
        print(f"❌ Aptive file not found: {aptive_file}")
        return
        
    if not os.path.exists(bom_template):
        print(f"❌ BOM Template file not found: {bom_template}")
        return
    
    # Process the files
    try:
        output_file = process_aptive_to_bom(aptive_file, bom_template)
        print(f"\n🎉 Success! Check your output file: {os.path.basename(output_file)}")
        
    except Exception as e:
        print(f"\n💥 Failed to process files: {str(e)}")

if __name__ == "__main__":
    main()