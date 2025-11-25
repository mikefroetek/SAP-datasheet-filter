"""
Optimized Universal Excel Level Processor
Handles unlimited levels (1 to 1000+) with efficient processing.
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os
import re
from datetime import datetime

def process_aptive_to_bom_optimized(aptive_file_path, bom_template_path, output_path=None):
    """
    Process Aptive Excel file with unlimited levels - optimized version.
    """
    
    print("🔄 Starting Optimized Universal Excel processing...")
    
    # Generate output path if not provided
    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_path = os.path.dirname(bom_template_path)
        output_path = os.path.join(base_path, f"BOM_Optimized_{timestamp}.xlsx")
    
    try:
        # Step 1: Read Aptive file
        print(f"📖 Reading Aptive file: {os.path.basename(aptive_file_path)}")
        
        if aptive_file_path.endswith('.xlsb'):
            df = pd.read_excel(aptive_file_path, engine='pyxlsb')
        else:
            df = pd.read_excel(aptive_file_path)
        
        print(f"✓ Read {len(df)} rows from Aptive file")
        
        # Analyze level distribution
        levels_found = set()
        for i in range(1, len(df)):
            if not pd.isna(df.iloc[i, 0]):
                levels_found.add(int(df.iloc[i, 0]))
        
        max_level = max(levels_found) if levels_found else 1
        print(f"📊 Levels detected: {sorted(levels_found)} (Max: {max_level})")
        
        # Step 2: Load BOM template
        workbook = load_workbook(bom_template_path)
        worksheet = workbook.active
        current_row = 9  # Start from row 9
        
        # Step 3: Build hierarchy map first
        print("🔄 Building hierarchy map...")
        hierarchy_map = build_hierarchy_map(df)
        
        # Step 4: Process hierarchies
        print("✏️ Processing hierarchies...")
        processed_groups = 0
        
        for level1_item in hierarchy_map:
            processed_groups += 1
            current_row = write_hierarchy_to_bom(
                hierarchy_map[level1_item], worksheet, current_row, processed_groups
            )
        
        # Save output
        workbook.save(output_path)
        
        print(f"✅ Processing completed successfully!")
        print(f"📊 Processed {processed_groups} hierarchy groups")
        print(f"📁 Output saved to: {output_path}")
        print(f"🚀 Handled levels 1 to {max_level}")
        
        return output_path
        
    except Exception as e:
        print(f"❌ Error during processing: {str(e)}")
        raise

def build_hierarchy_map(df):
    """
    Build a hierarchical map of all items and their relationships
    """
    items = {}
    
    # First pass: collect all items
    for i in range(1, len(df)):
        if pd.isna(df.iloc[i, 0]):
            continue
            
        level = int(df.iloc[i, 0])
        art_nr = str(int(df.iloc[i, 1])) if not pd.isna(df.iloc[i, 1]) else ""
        unit = str(df.iloc[i, 3]).upper() if not pd.isna(df.iloc[i, 3]) else ""
        quantity = df.iloc[i, 4] if not pd.isna(df.iloc[i, 4]) else ""
        
        items[i] = {
            'level': level,
            'art_nr': art_nr,
            'unit': unit,
            'quantity': quantity,
            'children': []
        }
    
    # Second pass: build parent-child relationships
    item_indices = list(items.keys())
    for i, idx in enumerate(item_indices):
        item = items[idx]
        current_level = item['level']
        
        # Look for children (next level down)
        for j in range(i + 1, len(item_indices)):
            child_idx = item_indices[j]
            child_item = items[child_idx]
            child_level = child_item['level']
            
            if child_level == current_level + 1:
                # Direct child
                item['children'].append(child_idx)
            elif child_level <= current_level:
                # Same or higher level, stop looking
                break
    
    # Third pass: organize by Level 1 items
    level1_hierarchies = {}
    for idx, item in items.items():
        if item['level'] == 1:
            level1_hierarchies[idx] = build_complete_hierarchy(items, idx)
    
    return level1_hierarchies

def build_complete_hierarchy(items, root_idx):
    """
    Build complete hierarchy tree starting from root
    """
    root_item = items[root_idx].copy()
    
    # Recursively build children
    complete_children = []
    for child_idx in root_item['children']:
        complete_children.append(build_complete_hierarchy(items, child_idx))
    
    root_item['complete_children'] = complete_children
    root_item['index'] = root_idx
    
    return root_item

def write_hierarchy_to_bom(hierarchy, worksheet, current_row, group_num):
    """
    Write hierarchy to BOM worksheet recursively
    """
    art_nr = hierarchy['art_nr']
    level = hierarchy['level']
    
    indent = "  " * (level - 1)
    print(f"{indent}📦 Group {group_num} Level {level}: {art_nr}")
    
    # Write current item as Material (Component empty)
    worksheet.cell(row=current_row, column=5, value=art_nr)
    worksheet.cell(row=current_row, column=14, value="")
    worksheet.cell(row=current_row, column=15, value="")
    worksheet.cell(row=current_row, column=16, value="")
    current_row += 1
    
    # Write all direct children as Components under this Material
    for child in hierarchy['complete_children']:
        child_art_nr = child['art_nr']
        child_unit = child['unit']
        child_quantity = child['quantity']
        
        indent_child = "  " * level
        print(f"{indent_child}🔗 Level {child['level']} under {art_nr}: {child_art_nr}")
        
        # Child as Component under current Material
        worksheet.cell(row=current_row, column=5, value=art_nr)
        worksheet.cell(row=current_row, column=14, value=child_art_nr)
        worksheet.cell(row=current_row, column=15, value=child_quantity)
        worksheet.cell(row=current_row, column=16, value=child_unit)
        current_row += 1
    
    # Now each child that has its own children becomes a Material
    for child in hierarchy['complete_children']:
        if child['complete_children']:  # Has grandchildren
            current_row = write_hierarchy_to_bom(child, worksheet, current_row, group_num)
    
    return current_row

def main():
    """Main function"""
    
    # File paths
    aptive_file = r"d:\SAP\Aptive_BOOM_v01.xlsb"
    bom_template = r"d:\SAP\BOM_Template.xlsx"
    
    # Check if files exist
    if not os.path.exists(aptive_file):
        print(f"❌ Aptive file not found: {aptive_file}")
        return
        
    if not os.path.exists(bom_template):
        print(f"❌ BOM Template file not found: {bom_template}")
        return
    
    # Process the files
    try:
        output_file = process_aptive_to_bom_optimized(aptive_file, bom_template)
        print(f"\n🎉 Success! Optimized universal processor!")
        print(f"📁 Check your output file: {os.path.basename(output_file)}")
        
    except Exception as e:
        print(f"\n💥 Failed to process files: {str(e)}")

if __name__ == "__main__":
    main()